from openpyxl import load_workbook

wb = load_workbook('data.xlsx')

ws = wb['Form responses 1']

max_rows= ws.max_row - 1

stateNames = ['Punjab','Bihar','Chandigarh','Chhattisgarh','Delhi','Gujarat','Haryana','Himachal Pradesh','India','Jharkhand','Kerala','Madhya Pradesh','Odisha','Rajasthan','Telangana','Uttar Pradesh','Uttarakhand','Maharashtra']
wrongState = [['punjab','amritsar'],['bihar','बिहार'],['c.g.','cg','chandigarh'],['chattisgarh','chhatisgarh','chhattisgarh','chhattishgarh','chhttisgarh','mungeli','छत्तीसगढ़','छत्तीसगढ़'],['dehli','delhi','ds','new delhi','rohini','4th'],['gujarat','gujrat','ગુજરાત'],['faridabad','harayana','hariyana','haryana','haryna','hriyana','hryana','kathal','matloda','mota loda','samlehri','हरियाणा'],['himachal pardesh','himachal pradesh','hp','h.p'],['india','indian'],['jharkhand'],['karale','kerala','keralam','kerela','vengara nedumpambu'],['m p','m. p.','m.p','m.p.','barwani','madh8','madhay pradesh','madhy pradesh','madhya pardesh','madhya pradesh','madhya prdesh','madhyapradesh','mp','खरगोन','मथ्यप्रदेश','मध्य प्रदेश','मध्यप्रदेश'],['odisha','orissa'],['rajasthan','rajasthana'],['telangaana','telangan','telangana','telangana state','telanganna'],['u p','u. p','u. p.','u.p','u.p.','up','sbi','uterperdesh','uttar pardesh','uttar peadesh','uttar pradesh','uttar prdesh','barabanki','uttara pradesh','uttarpardesh','uttarpradesh','utter  perdesh','utter pardesh','utter pradesh','utter-pradesh','उत्तर प्रदेश','उत्तरप्रदेश','उत्‍तर प्रदेश'],['uttarakhand','uttrakhand'],['महाराष्ट्र']]

def deleteDuplicate(originalArr,modifiedArr):
  modifiedArr.append(originalArr[0])
  for i in originalArr:
    exists = False

    for j in modifiedArr:
      if(i == j):
        exists=True
        break
      else:
        continue
    if(not exists):
      modifiedArr.append(i)

states = []
lc_states=[]
diff_states=[]
for i in range(max_rows):
  
  for idx, j in enumerate(wrongState):
    for k in j:
      if(ws.cell(row=i+2,column=8).value):
        exists=False
        for l in stateNames:
          if(ws.cell(row=i+2,column=8).value != l):
            exists=True
        
        if(not exists):
          my_file = open("differentStateList.txt", "w" ,encoding="utf-8")
          my_file.write(ws.cell(row=i+2,column=8).value)
          my_file.close()
        # if((ws.cell(row=i+2,column=8).value).lower().strip() == k):
        #   ws.cell(row=i+2,column=8).value= stateNames[idx]
          
        


# for i in states:
#   lc_states.append(i.lower().strip())

# lc_states.sort()

# diff_states.append(lc_states[0])
# for i in lc_states:
#   exists = False

#   for j in diff_states:
#     if(i == j):
#       exists=True
#       break
#     else:
#       continue
#   if(not exists):
#     diff_states.append(i)

# print(diff_states)
# CorrectedStates =[]
# AllCorrectedStates=[]

# for i in lc_states:
#   found = False
#   for idx, j in enumerate(wrongState):
#     for k in j:
#       if(i == k):
#         AllCorrectedStates.append(stateNames[idx])
#         found=True

# deleteDuplicate(AllCorrectedStates,CorrectedStates)

# print(CorrectedStates)
        
      


# stringStates = '['
# for i in diff_states:
#   stringStates+=i+',\n'


# my_file = open("differentStateList.txt", "w" ,encoding="utf-8")
# my_file.write(stringStates)
# my_file.close()

wb.save("data.xlsx") 